from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import CustomUserCreationForm, EditarUsuarioForm, CargaFamiliarForm, SolicitudForm, ContactForm, EditProfilePhotoForm, PasswordChangeForm, CursoForm, EditarParticipantesForm, BeneficioForm, DenunciaForm, EvidenciaFormset, NotaDenunciaForm, PublicacionForm, CustomUserChangeForm, DocumentoEmpresaForm, SolicitudVacacionesForm, CustomPasswordChangeForm,ActualizarProgresoForm
import os
from django.conf import settings
from .models import CustomUser, CargaFamiliar, Asistencia, Solicitud, Curso, Beneficio, Area, Denuncia, EvidenciaDenuncia, DocumentoEmpresa, SolicitudVacaciones, DescargaDocumento, ProgresoParticipante
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template
import openpyxl
from datetime import datetime, timedelta
from django.urls import reverse
from django.core.mail import send_mail
from django.db.models import Q
from django.contrib.auth.models import Group
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from openpyxl import Workbook
from django.template.defaultfilters import register
from django import template

def is_hr_analyst(user):
    return user.area and user.area.nombre == 'Recursos Humanos' and user.cargo == 'Analista de Personas'

#Función para separar usuarios por grupo supervisores o colaboradores
def is_supervisor(user):
    return user.groups.filter(name='supervisores').exists()

@login_required
def estadisticas_rrhh(request):
    # Cantidad de usuarios excluyendo los que comienzan con "admin"
    cantidad_usuarios = CustomUser.objects.exclude(username__startswith='admin').count()

    # Cantidad de solicitudes por tipo y estado
    solicitudes_por_tipo_estado = (
        Solicitud.objects.values('tipo', 'estado')
        .annotate(cantidad=Count('id'))
        .order_by('tipo', 'estado')
    )

    # Cantidad de solicitudes de vacaciones por estado
    solicitudes_vacaciones_por_estado = (
        SolicitudVacaciones.objects.values('estado')
        .annotate(cantidad=Count('id'))
    )

    # Cantidad de denuncias por estado
    denuncias_por_estado = (
        Denuncia.objects.values('estado')
        .annotate(cantidad=Count('id'))
    )

    # Cantidad de cursos
    cantidad_cursos = Curso.objects.count()

    # Total de denuncias
    cantidad_denuncias = Denuncia.objects.count()

    context = {
        'cantidad_usuarios': cantidad_usuarios,
        'solicitudes_por_tipo_estado': solicitudes_por_tipo_estado,
        'solicitudes_vacaciones_por_estado': solicitudes_vacaciones_por_estado,
        'denuncias_por_estado': denuncias_por_estado,
        'cantidad_cursos': cantidad_cursos,
        'cantidad_denuncias': cantidad_denuncias,
    }

    return render(request, 'home.html', context)

#Vista para buscar
@login_required
def buscar(request):
    query = request.GET.get('q')

    # Iniciar el contexto con los resultados vacíos
    context = {
        'resultados_solicitudes': [],
        'resultados_cursos': [],
        'resultados_usuarios': [],
        'resultados_vacaciones': [],
        'resultados_beneficios': [],
        'resultados_denuncias': [],
        'resultados_documentos': [],
    }

    if not query:
        return render(request, 'gestiones/buscar/busqueda.html', context)

    if request.user.groups.filter(name='supervisores').exists():
        # Supervisores
        context['resultados_solicitudes'] = Solicitud.objects.filter(
            Q(tipo__icontains=query) | Q(descripcion__icontains=query) | Q(estado__icontains=query),
            colaborador__empresa=request.user.empresa
        )
        context['resultados_cursos'] = Curso.objects.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query),
            supervisor__empresa=request.user.empresa
        ).distinct()
        context['resultados_usuarios'] = CustomUser.objects.filter(
            Q(username__icontains=query) | Q(first_name__icontains=query) | Q(last_name__icontains=query),
            empresa=request.user.empresa
        )
        context['resultados_vacaciones'] = SolicitudVacaciones.objects.filter(
            Q(motivo__icontains=query) | Q(estado__icontains=query),
            colaborador__empresa=request.user.empresa
        )
        context['resultados_beneficios'] = Beneficio.objects.filter(
            Q(titulo__icontains=query) | Q(descripcion__icontains=query),
            creado_por__empresa=request.user.empresa
        )
        context['resultados_denuncias'] = Denuncia.objects.filter(
            Q(motivo__icontains=query) | Q(descripcion__icontains=query),
            denunciante__empresa=request.user.empresa
        )
        context['resultados_documentos'] = DocumentoEmpresa.objects.filter(
            Q(titulo__icontains=query) | Q(descripcion__icontains=query),
            creado_por__empresa=request.user.empresa
        )
    else:
        # Colaboradores
        context['resultados_solicitudes'] = Solicitud.objects.filter(
            Q(tipo__icontains=query) | Q(descripcion__icontains=query) | Q(estado__icontains=query),
            colaborador=request.user
        )
        context['resultados_cursos'] = Curso.objects.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query),
            participantes=request.user
        ).distinct()
        context['resultados_vacaciones'] = SolicitudVacaciones.objects.filter(
            Q(motivo__icontains=query) | Q(estado__icontains=query),
            colaborador=request.user
        )
        context['resultados_beneficios'] = Beneficio.objects.filter(
            Q(titulo__icontains=query) | Q(descripcion__icontains=query),
            creado_por__empresa=request.user.empresa
        )
        context['resultados_documentos'] = DocumentoEmpresa.objects.filter(
            Q(titulo__icontains=query) | Q(descripcion__icontains=query),
            creado_por__empresa=request.user.empresa
        )

    return render(request, 'gestiones/buscar/busqueda.html', context)
@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())  # Solo supervisores pueden acceder
def registrar_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            new_user = form.save(commit=False)
            new_user.empresa = request.user.empresa  # Asigna la empresa del usuario que registra
            new_user.set_password(form.cleaned_data['password1'])  # Asegúrate de configurar la contraseña
            new_user.save()

            # Asignar grupo al usuario
            grupo = form.cleaned_data.get('grupo')
            if grupo:
                new_user.groups.add(grupo)

            messages.success(request, "Usuario registrado correctamente.")
            return redirect('lista_colaboradores')  # Cambia esta URL según tu proyecto
        else:
            messages.error(request, "Corrige los errores en el formulario.")
    else:
        form = CustomUserCreationForm(user=request.user)

    return render(request, 'gestiones/usuarios/registrar_usuario.html', {'form': form})

from datetime import date

@login_required
@user_passes_test(is_supervisor)
def ficha_usuario(request, pk):
    usuario = get_object_or_404(CustomUser, pk=pk)

    # Verificar que el usuario pertenece a la misma empresa que el supervisor
    if usuario.empresa != request.user.empresa:
        messages.error(request, "No tienes permiso para ver esta ficha.")
        return redirect('lista_colaboradores')

    # Calcular la edad si la fecha de nacimiento está definida
    if usuario.fecha_nacimiento:
        today = date.today()
        edad = today.year - usuario.fecha_nacimiento.year - (
            (today.month, today.day) < (usuario.fecha_nacimiento.month, usuario.fecha_nacimiento.day)
        )
    else:
        edad = None

    # Obtener los documentos descargados por el usuario
    descargas = DescargaDocumento.objects.filter(usuario=usuario)

    # Pasar documentos al contexto
    return render(request, 'gestiones/usuarios/ficha_usuario.html', {
        'usuario': usuario,
        'edad': edad,  # Pasar la edad al contexto
        'descargas': descargas,  # Pasar los documentos descargados al contexto
    })

@login_required
@user_passes_test(is_supervisor)
def ficha_usuario_pdf(request, pk):
    usuario = get_object_or_404(CustomUser, pk=pk)

    # Verificar que el usuario pertenece a la misma empresa que el supervisor
    if usuario.empresa != request.user.empresa:
        messages.error(request, "No tienes permiso para ver esta ficha.")
        return redirect('lista_colaboradores')

    # Calcular la edad si la fecha de nacimiento está definida
    if usuario.fecha_nacimiento:
        today = date.today()
        edad = today.year - usuario.fecha_nacimiento.year - (
            (today.month, today.day) < (usuario.fecha_nacimiento.month, usuario.fecha_nacimiento.day)
        )
    else:
        edad = None

    # Renderizar la plantilla a HTML
    template_path = 'gestiones/usuarios/ficha_usuario_pdf.html'
    context = {'usuario': usuario, 'edad': edad}
    template = get_template(template_path)
    html = template.render(context)

    # Crear un archivo PDF desde el HTML
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=ficha_usuario_{usuario.username}.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Manejar errores
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF.')
    return response

@login_required
@user_passes_test(is_supervisor)
def lista_colaboradores(request):
    colaboradores = CustomUser.objects.filter(empresa=request.user.empresa).order_by('date_joined')  # Orden ascendente por fecha de registro
    return render(request, 'gestiones/usuarios/lista_colaboradores.html', {'colaboradores': colaboradores})


@login_required
@user_passes_test(is_supervisor)
def editar_colaborador(request, pk):
    colaborador = get_object_or_404(CustomUser, pk=pk)
    
    if request.method == 'POST':
        # Asegúrate de pasar request.FILES al formulario
        form = CustomUserChangeForm(request.POST, request.FILES, instance=colaborador)
        if form.is_valid():
            # Guarda los datos del formulario (incluidos los archivos)
            colaborador = form.save(commit=False)

            # Maneja archivos individualmente si no son procesados automáticamente
            if 'certificado_afp' in request.FILES:
                colaborador.certificado_afp = request.FILES['certificado_afp']
            if 'certificado_salud' in request.FILES:
                colaborador.certificado_salud = request.FILES['certificado_salud']

            colaborador.save()

            # Maneja los grupos del colaborador
            grupos = form.cleaned_data['grupo']
            colaborador.groups.set(grupos)

            messages.success(request, 'El colaborador ha sido actualizado correctamente.')
            return redirect('lista_colaboradores')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = CustomUserChangeForm(instance=colaborador)
    
    return render(request, 'gestiones/usuarios/editar_colaborador.html', {'form': form, 'colaborador': colaborador})




@login_required
@user_passes_test(is_supervisor)
def eliminar_colaborador(request, pk):
    colaborador = get_object_or_404(CustomUser, pk=pk)
    
    if request.method == 'POST':
        colaborador.delete()
        return redirect('lista_colaboradores')
    
    return render(request, 'gestiones/usuarios/eliminar_colaborador.html', {'colaborador': colaborador})

@login_required
@user_passes_test(is_supervisor)
def exportar_colaboradores_excel(request):
    colaboradores = CustomUser.objects.filter(empresa=request.user.empresa)
    
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Colaboradores'

    # Escribir encabezados
    ws.append(['Username', 'Nombre', 'Apellido', 'Email', 'Cargo'])

    # Escribir datos de colaboradores
    for colaborador in colaboradores:
        ws.append([colaborador.username, colaborador.first_name, colaborador.last_name, colaborador.email, colaborador.cargo])
    
    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=colaboradores.xlsx'
    wb.save(response)
    
    return response


#Vistas Cargas Familiares
@login_required
@user_passes_test(is_supervisor)
def registrar_carga(request):
    if request.method == 'POST':
        form = CargaFamiliarForm(request.POST, request.FILES, user=request.user) 
        if form.is_valid():
            form.save()
            return redirect('listar_cargas')
    else:
        form = CargaFamiliarForm(user=request.user)  # Pasa el usuario al formulario
    return render(request, 'gestiones/cargas_familiares/registrar_carga.html', {'form': form})


@login_required
def listar_cargas(request):
    if is_supervisor(request.user):
        # Si el usuario es supervisor, muestra todas las cargas de su empresa
        cargas = CargaFamiliar.objects.filter(usuario__empresa=request.user.empresa)

        # Obtener la lista de usuarios para el filtro
        usuarios = CustomUser.objects.filter(empresa=request.user.empresa)
    else:
        # Si el usuario es colaborador, muestra solo sus cargas
        cargas = CargaFamiliar.objects.filter(usuario=request.user)
        usuarios = []  # Los colaboradores no necesitan ver este filtro

    # Obtener parámetros de filtrado desde el formulario
    nombre = request.GET.get('nombre')
    usuario_id = request.GET.get('usuario_id')  # Filtro de usuario

    # Aplicar filtros si están presentes
    if nombre:
        cargas = cargas.filter(nombre__icontains=nombre)
    if usuario_id:
        cargas = cargas.filter(usuario_id=usuario_id)

    return render(request, 'gestiones/cargas_familiares/listar_cargas.html', {
        'cargas': cargas,
        'nombre': nombre,  # Para mantener el valor en el formulario
        'usuario_id': usuario_id,  # Para mantener el filtro seleccionado
        'usuarios': usuarios,  # Lista de usuarios para el filtro
    })


@login_required
def editar_carga(request, pk):
    carga = get_object_or_404(CargaFamiliar, pk=pk)
    if request.method == 'POST':
        form = CargaFamiliarForm(request.POST, instance=carga)
        if form.is_valid():
            form.save()
            return redirect('listar_cargas')
    else:
        form = CargaFamiliarForm(instance=carga)
    return render(request, 'gestiones/cargas_familiares/editar_carga.html', {'form': form})

@login_required
def eliminar_carga(request, pk):
    carga = get_object_or_404(CargaFamiliar, pk=pk)
    if request.method == 'POST':
        carga.delete()
        return redirect('listar_cargas')
    return render(request, 'gestiones/cargas_familiares/eliminar_carga.html', {'carga': carga})

#Vistas Asistencia
@login_required
def registro_asistencia(request):
    asistencia = Asistencia.objects.filter(colaborador=request.user, fecha=datetime.now().date()).first()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if not asistencia:
            asistencia = Asistencia(colaborador=request.user)

        current_time = datetime.now()

        if action == 'entrada':
            asistencia.hora_entrada = current_time.time()
        elif action == 'salida':
            asistencia.hora_salida = current_time.time()

        asistencia.save()
        return redirect('visualizacion_asistencia')
    
    tiempo_trabajado = None
    if asistencia and asistencia.hora_entrada and asistencia.hora_salida:
        entrada = datetime.combine(asistencia.fecha, asistencia.hora_entrada)
        salida = datetime.combine(asistencia.fecha, asistencia.hora_salida)
        tiempo_trabajado_delta = salida - entrada
        
        # Formatear a HH:MM:SS
        total_segundos = int(tiempo_trabajado_delta.total_seconds())
        horas, resto = divmod(total_segundos, 3600)
        minutos, segundos = divmod(resto, 60)
        tiempo_trabajado = f"{horas:02}:{minutos:02}:{segundos:02}"
    
    return render(request, 'gestiones/asistencia/registro_asistencia.html', {
        'asistencia': asistencia,
        'tiempo_trabajado': tiempo_trabajado
    })

@login_required
def visualizacion_asistencia(request):

    usuario_id = None

    # Verificar si el usuario es un supervisor o un colaborador regular
    if request.user.area and request.user.area.nombre == 'Recursos Humanos' and request.user.groups.filter(name='supervisores').exists():
        # Supervisores ven todas las asistencias de la empresa
        asistencias = Asistencia.objects.filter(colaborador__empresa=request.user.empresa)
        
        # Filtrar por usuario si el parámetro usuario_id está presente
        usuario_id = request.GET.get('usuario_id')
        if usuario_id:
            asistencias = asistencias.filter(colaborador_id=usuario_id)
    else:
        # Colaboradores regulares solo ven sus propias asistencias
        asistencias = Asistencia.objects.filter(colaborador=request.user)

    # Obtener los parámetros de filtro de fecha desde el formulario
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio and fecha_fin:
        try:
            # Asegurarse de que las fechas estén en el formato correcto
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            asistencias = asistencias.filter(fecha__range=[fecha_inicio, fecha_fin])
        except ValueError:
            # Manejar el error si las fechas no están en el formato esperado
            pass

    # Procesar los datos para calcular tiempos trabajados y faltantes
    datos_asistencias = []
    for asistencia in asistencias:
        tiempo_trabajado = None
        tiempo_faltante = None
        if asistencia.hora_entrada and asistencia.hora_salida:
            entrada = datetime.combine(asistencia.fecha, asistencia.hora_entrada)
            salida = datetime.combine(asistencia.fecha, asistencia.hora_salida)
            tiempo_trabajado = salida - entrada
            tiempo_planificado = timedelta(hours=8)  # Jornada estándar de 8 horas
            tiempo_faltante = tiempo_planificado - tiempo_trabajado

        datos_asistencias.append({
            'fecha': asistencia.fecha,
            'hora_entrada': asistencia.hora_entrada,
            'hora_salida': asistencia.hora_salida,
            'tiempo_trabajado': tiempo_trabajado,
            'tiempo_faltante': tiempo_faltante,
            'colaborador': asistencia.colaborador.username,
        })

    # Obtener lista de usuarios (solo para supervisores)
    colaboradores = []
    if request.user.groups.filter(name='supervisores').exists():
        colaboradores = CustomUser.objects.filter(empresa=request.user.empresa)

    return render(request, 'gestiones/asistencia/visualizacion_asistencia.html', {
        'datos_asistencias': datos_asistencias,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'colaboradores': colaboradores,  # Pasar lista de colaboradores al template
        'usuario_id': usuario_id,  # Mantener el filtro seleccionado en el formulario
    })




@login_required
@user_passes_test(is_supervisor)  # Asegúrate de usar un decorador adecuado según los permisos
def exportar_asistencias_excel(request):
    # Filtrar asistencias por la empresa del usuario y las fechas (si se pasan como parámetros)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Obtener las asistencias según los filtros
    if fecha_inicio and fecha_fin:
        asistencias = Asistencia.objects.filter(
            colaborador__empresa=request.user.empresa,
            fecha__range=[fecha_inicio, fecha_fin]
        )
    else:
        asistencias = Asistencia.objects.filter(colaborador__empresa=request.user.empresa)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencias'

    # Escribir encabezados
    ws.append(['Fecha', 'Hora de Entrada', 'Hora de Salida', 'Colaborador', 'Tiempo Trabajado'])

    # Escribir datos de asistencias
    for asistencia in asistencias:
        tiempo_trabajado = None
        if asistencia.hora_entrada and asistencia.hora_salida:
            entrada = datetime.combine(asistencia.fecha, asistencia.hora_entrada)
            salida = datetime.combine(asistencia.fecha, asistencia.hora_salida)
            tiempo_trabajado = salida - entrada

        ws.append([
            asistencia.fecha,
            asistencia.hora_entrada if asistencia.hora_entrada else 'N/A',
            asistencia.hora_salida if asistencia.hora_salida else 'N/A',
            asistencia.colaborador.username,
            str(tiempo_trabajado) if tiempo_trabajado else 'N/A'
        ])

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asistencias.xlsx'
    wb.save(response)

    return response

@login_required
def exportar_asistencias_excel(request):
    # Filtrar asistencias por la empresa del usuario y las fechas (si se pasan como parámetros)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Obtener las asistencias según los filtros
    if fecha_inicio and fecha_fin:
        asistencias = Asistencia.objects.filter(
            colaborador__empresa=request.user.empresa,
            fecha__range=[fecha_inicio, fecha_fin]
        )
    else:
        asistencias = Asistencia.objects.filter(colaborador__empresa=request.user.empresa)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencias'

    # Escribir encabezados
    ws.append(['Fecha', 'Hora de Entrada', 'Hora de Salida', 'Colaborador', 'Tiempo Trabajado'])

    # Escribir datos de asistencias
    for asistencia in asistencias:
        tiempo_trabajado = None
        if asistencia.hora_entrada and asistencia.hora_salida:
            entrada = datetime.combine(asistencia.fecha, asistencia.hora_entrada)
            salida = datetime.combine(asistencia.fecha, asistencia.hora_salida)
            tiempo_trabajado = salida - entrada

        ws.append([
            asistencia.fecha,
            asistencia.hora_entrada if asistencia.hora_entrada else 'N/A',
            asistencia.hora_salida if asistencia.hora_salida else 'N/A',
            asistencia.colaborador.username,
            str(tiempo_trabajado) if tiempo_trabajado else 'N/A'
        ])

    # Configurar la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asistencias.xlsx'
    wb.save(response)

    return response

#Solicitudes Vacaciones

@login_required
def crear_solicitud_vacaciones(request):
    user = request.user
    dias_disponibles = user.dias_vacaciones_disponibles

    # Restricción: No permitir solicitudes si los días disponibles son 0
    if dias_disponibles <= 0:
        messages.error(request, 'No puedes solicitar más días de vacaciones. No tienes días disponibles.')
        return redirect('lista_solicitudes_vacaciones')

    if request.method == 'POST':
        form = SolicitudVacacionesForm(request.POST, user=user)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.colaborador = user
            solicitud.save()
            messages.success(request, 'Solicitud de vacaciones creada correctamente.')
            return redirect('lista_solicitudes_vacaciones')
    else:
        form = SolicitudVacacionesForm(user=user)

    return render(request, 'gestiones/vacaciones/crear_solicitud_vacaciones.html', {
        'form': form,
        'dias_disponibles': dias_disponibles,
    })

@login_required
def gestionar_solicitud_vacaciones(request, pk):
    solicitud = get_object_or_404(SolicitudVacaciones, pk=pk)
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        if accion == 'aprobar':
            solicitud.estado = 'aprobada'
        elif accion == 'rechazar':
            solicitud.estado = 'rechazada'
        solicitud.save()
        return redirect('lista_solicitudes_vacaciones')
    
    return render(request, 'gestiones/vacaciones/gestionar_solicitud_vacaciones.html', {'solicitud': solicitud})


@login_required
def lista_solicitudes_vacaciones(request):
    """
    Vista para listar solicitudes de vacaciones.
    Los supervisores ven las solicitudes de su área, y los colaboradores ven sus propias solicitudes.
    """
    if request.user.groups.filter(name='supervisores').exists():
        # Si el usuario es supervisor, muestra las solicitudes de su área
        solicitudes = SolicitudVacaciones.objects.filter(
            colaborador__area=request.user.area
        ).order_by('-fecha_creacion')
        es_supervisor = True
    else:
        # Si el usuario es colaborador, muestra solo sus propias solicitudes
        solicitudes = SolicitudVacaciones.objects.filter(
            colaborador=request.user
        ).order_by('-fecha_creacion')
        es_supervisor = False

    return render(request, 'gestiones/vacaciones/lista_solicitudes_vacaciones.html', {
        'solicitudes': solicitudes,
        'es_supervisor': es_supervisor
    })

#Vistas Solicitudes
@login_required
def crear_solicitud(request):
    if request.method == 'POST':
        form = SolicitudForm(request.POST)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.colaborador = request.user
            solicitud.save()
            return redirect('lista_solicitudes')
    else:
        form = SolicitudForm()
    return render(request, 'gestiones/solicitudes/crear_solicitud.html', {'form': form})

@login_required
def lista_solicitudes(request):
    if request.user.groups.filter(name='supervisores').exists():
        if request.user.area and request.user.area.nombre == 'Recursos Humanos':
            # Supervisores del área de Recursos Humanos pueden ver todas las solicitudes
            solicitudes = Solicitud.objects.all()
        else:
            # Otros supervisores solo pueden ver las solicitudes de su área
            solicitudes = Solicitud.objects.filter(colaborador__area=request.user.area)
        is_supervisor = True
    else:
        # Los demás usuarios solo pueden ver sus propias solicitudes
        solicitudes = Solicitud.objects.filter(colaborador=request.user)
        is_supervisor = False

    # Agregamos la bandera es_supervisor al contexto
    context = {
        'solicitudes': solicitudes,
        'is_supervisor': is_supervisor
    }

    return render(request, 'gestiones/solicitudes/lista_solicitudes.html', context)

@login_required
def gestionar_solicitud(request, pk):
    solicitud = get_object_or_404(Solicitud, pk=pk)
    if request.method == 'POST':
        if 'aprobar' in request.POST:
            solicitud.estado = 'aprobada'
        elif 'rechazar' in request.POST:
            solicitud.estado = 'rechazada'
        solicitud.save()
        return redirect('lista_solicitudes')
    return render(request, 'gestiones/solicitudes/gestionar_solicitud.html', {'solicitud': solicitud})

@login_required
def descargar_comprobante(request, pk):
    solicitud = get_object_or_404(Solicitud, pk=pk)
    template_path = 'gestiones/solicitudes/comprobante_solicitud.html'
    context = {'solicitud': solicitud}
    
    # Cargar la plantilla y renderizarla con el contexto
    template = get_template(template_path)
    html = template.render(context)
    
    # Crear un objeto de respuesta como PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Comprobante_Solicitud_{solicitud.id}.pdf"'
    
    # Generar el PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    # Verificar si hubo errores
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF: <pre>' + html + '</pre>')
    
    return response
#Vista Contacto
def contacto(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            motivo_contacto = form.cleaned_data['motivo_contacto']
            nombre = form.cleaned_data['nombre']
            apellido = form.cleaned_data['apellido']
            numero_telefono = form.cleaned_data['numero_telefono']
            correo = form.cleaned_data['correo']
            nombre_empresa = form.cleaned_data['nombre_empresa']
            cantidad_colaboradores = form.cleaned_data['cantidad_colaboradores']
            cargo = form.cleaned_data['cargo']
            area_desempeno = form.cleaned_data['area_desempeno']
            rubro = form.cleaned_data['rubro']
            mensaje = form.cleaned_data['mensaje']
            
            send_mail(
                f'Contacto desde la web: {motivo_contacto}',
                f'Nombre: {nombre} {apellido}\nTeléfono: {numero_telefono}\nCorreo: {correo}\nEmpresa: {nombre_empresa}\nColaboradores: {cantidad_colaboradores}\nCargo: {cargo}\nÁrea de Desempeño: {area_desempeno}\nRubro: {rubro}\nMensaje: {mensaje}',
                'tu_email@gmail.com',
                ['destino_email@gmail.com'],
                fail_silently=False,
            )
            
            return redirect('success')
    else:
        form = ContactForm()
    
    return render(request, 'gestiones/contacto/contacto.html', {'form': form})

def success(request):
    return render(request, 'gestiones/contacto/success.html')

# Vista del perfil de usuario
@login_required
def profile(request):
    user = request.user
    edad = None
    if user.fecha_nacimiento:
        today = date.today()
        edad = today.year - user.fecha_nacimiento.year - ((today.month, today.day) < (user.fecha_nacimiento.month, user.fecha_nacimiento.day))
    
    return render(request, 'gestiones/profile/profile.html', {
        'user': user,
        'edad': edad,
        'dias_vacaciones': user.dias_vacaciones_disponibles,
    })

# Vista para editar la foto de perfil
@login_required
def edit_profile_photo(request):
    if request.method == 'POST':
        form = EditProfilePhotoForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = EditProfilePhotoForm(instance=request.user)
    
    return render(request, 'gestiones/profile/edit_profile_photo.html', {'form': form})

# Vista para cambiar la contraseña
@login_required
def cambiar_contrasena(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()  # Guarda la nueva contraseña
            update_session_auth_hash(request, user)  # Mantiene al usuario autenticado
            messages.success(request, "Tu contraseña ha sido cambiada exitosamente.")
            return redirect('profile')  # Redirige al perfil
        else:
            messages.error(request, "Por favor corrige los errores a continuación.")
    else:
        form = CustomPasswordChangeForm(user=request.user)

    return render(request, 'gestiones/profile/cambiar_contrasena.html', {'form': form})

#Vistas Cursos
@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def crear_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            curso = form.save(commit=False)
            curso.supervisor = request.user
            curso.save()
            return redirect('detalle_curso', curso.id)  # Redirigir al detalle del curso
    else:
        form = CursoForm()
    return render(request, 'gestiones/cursos/crear_curso.html', {'form': form})


@login_required
def detalle_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    progresos = curso.progresos.select_related('participante').all()  # Carga los progresos con los participantes

    return render(request, 'gestiones/cursos/detalle_curso.html', {
        'curso': curso,
        'progresos': progresos,  # Pasa los progresos al contexto
    })



@login_required
def lista_cursos(request):
    if request.user.groups.filter(name='supervisores').exists():
        # Si el usuario es supervisor, muestra todos los cursos de su empresa
        cursos = Curso.objects.filter(supervisor__empresa=request.user.empresa)
    else:
        # Si el usuario es colaborador, muestra solo los cursos en los que participa
        cursos = Curso.objects.filter(participantes=request.user)

    return render(request, 'gestiones/cursos/lista_cursos.html', {'cursos': cursos})



@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def editar_participantes(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id, supervisor=request.user)

    if request.method == 'POST':
        form = EditarParticipantesForm(request.POST, instance=curso)
        if form.is_valid():
            # Obtener participantes actuales antes de guardar
            participantes_anteriores = set(curso.participantes.all())

            # Guardar nuevos participantes
            form.save()

            # Obtener participantes actuales después de guardar
            participantes_actuales = set(curso.participantes.all())

            # Detectar participantes nuevos
            participantes_nuevos = participantes_actuales - participantes_anteriores

            # Crear progreso para los nuevos participantes
            for nuevo_participante in participantes_nuevos:
                ProgresoParticipante.objects.create(curso=curso, participante=nuevo_participante, progreso=0)

            # Eliminar progresos de los participantes eliminados
            participantes_eliminados = participantes_anteriores - participantes_actuales
            ProgresoParticipante.objects.filter(curso=curso, participante__in=participantes_eliminados).delete()

            return redirect('detalle_curso', curso_id=curso.id)
    else:
        form = EditarParticipantesForm(instance=curso)

    return render(request, 'gestiones/cursos/editar_participantes.html', {
        'curso': curso,
        'form': form,
    })

@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id, supervisor=request.user)

    if request.method == 'POST':
        curso.delete()
        messages.success(request, 'El curso ha sido eliminado correctamente.')
        return redirect('lista_cursos')  # Redirigir a la lista de cursos

    return render(request, 'gestiones/cursos/eliminar_curso.html', {'curso': curso})

@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def actualizar_progreso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id, supervisor=request.user)
    progresos = curso.progresos.select_related('participante').all()

    if request.method == 'POST':
        for progreso in progresos:
            nuevo_progreso = request.POST.get(f'progreso_{progreso.id}', None)
            if nuevo_progreso is not None:
                progreso.progreso = int(nuevo_progreso)
                progreso.save()
        return redirect('detalle_curso', curso_id=curso.id)

    return render(request, 'gestiones/cursos/actualizar_progreso.html', {
        'curso': curso,
        'progresos': progresos,
    })



#Vistas Beneficios
def lista_beneficios(request):
    empresa = request.user.empresa
    beneficios = Beneficio.objects.filter(creado_por__empresa=empresa)

    context = {
        'beneficios': beneficios
    }

    return render(request, 'gestiones/beneficios/lista_beneficios.html', context)

@login_required
def crear_beneficio(request):
    if request.method == 'POST':
        form = BeneficioForm(request.POST, request.FILES)
        if form.is_valid():
            beneficio = form.save(commit=False)
            beneficio.creado_por = request.user
            beneficio.save()
            return redirect('lista_beneficios')
    else:
        form = BeneficioForm()
    return render(request, 'gestiones/beneficios/crear_beneficio.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def editar_beneficio(request, id):
    beneficio = get_object_or_404(Beneficio, id=id)

    if request.method == 'POST':
        form = BeneficioForm(request.POST, request.FILES, instance=beneficio)
        if form.is_valid():
            form.save()
            return redirect('lista_beneficios')
    else:
        form = BeneficioForm(instance=beneficio)

    return render(request, 'gestiones/beneficios/editar_beneficio.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists())
def eliminar_beneficio(request, id):
    beneficio = get_object_or_404(Beneficio, id=id)

    if request.method == 'POST':
        beneficio.delete()
        return redirect('lista_beneficios')

    return render(request, 'gestiones/beneficios/eliminar_beneficio_confirmacion.html', {'beneficio': beneficio})



@login_required
def crear_denuncia(request):
    if request.method == 'POST':
        form = DenunciaForm(request.POST)
        formset = EvidenciaFormset(request.POST, request.FILES)
        
        if form.is_valid() and formset.is_valid():
            # Guardar la denuncia
            denuncia = form.save(commit=False)
            denuncia.denunciante = request.user  # Asignar el denunciante
            denuncia.save()

            # Guardar las evidencias relacionadas
            for evidencia_form in formset:
                if evidencia_form.cleaned_data.get('archivo'):  # Si se ha proporcionado un archivo
                    evidencia = evidencia_form.save(commit=False)
                    evidencia.denuncia = denuncia  # Relacionar con la denuncia
                    evidencia.save()

            return redirect('denuncia_creada')  # Redirige a una página de confirmación
    else:
        form = DenunciaForm()
        formset = EvidenciaFormset(queryset=EvidenciaDenuncia.objects.none())  # Formset vacío

    return render(request, 'gestiones/denuncias/crear_denuncia.html', {'form': form, 'formset': formset})

def denuncia_creada(request):
    return render(request, 'gestiones/denuncias/denuncia_creada.html')
                  
@login_required
def lista_denuncias(request):
    # Si el usuario es de Recursos Humanos y tiene el cargo de Analista de Personas
    if request.user.area.nombre == 'Recursos Humanos' and request.user.cargo == 'Analista de Personas':
        # Los usuarios de Recursos Humanos pueden ver todas las denuncias
        denuncias = Denuncia.objects.all()
    else:
        # Otros usuarios solo pueden ver las denuncias que ellos mismos han creado
        denuncias = Denuncia.objects.filter(denunciante=request.user)

    return render(request, 'gestiones/denuncias/lista_denuncias.html', {'denuncias': denuncias})

@login_required
def detalle_denuncia(request, pk):
    denuncia = get_object_or_404(Denuncia, pk=pk)
    notas = denuncia.notas.all()  # Notas relacionadas con la denuncia

    # Verificar si el usuario puede gestionar la denuncia
    es_analista_rrhh = request.user.area and request.user.area.nombre == 'Recursos Humanos' and request.user.cargo == 'Analista de Personas'

    if request.method == 'POST':
        form = NotaDenunciaForm(request.POST)
        if form.is_valid():
            nota = form.save(commit=False)  # No guardamos directamente aún
            nota.denuncia = denuncia        # Relacionamos la nota con la denuncia
            nota.usuario = request.user     # Establecemos el usuario que hizo la nota
            nota.save()                     # Ahora sí guardamos la nota
            return redirect('detalle_denuncia', pk=denuncia.pk)
    else:
        form = NotaDenunciaForm()

    return render(request, 'gestiones/denuncias/detalle_denuncia.html', {
        'denuncia': denuncia,
        'form': form,
        'notas': notas,
        'es_analista_rrhh': es_analista_rrhh
    })


@login_required
@user_passes_test(is_hr_analyst)
def gestionar_denuncia(request, pk):
    denuncia = get_object_or_404(Denuncia, pk=pk)
    
    if request.method == 'POST':
        if 'revision' in request.POST:
            denuncia.estado = 'revision'
        elif 'resuelta' in request.POST:
            denuncia.estado = 'resuelta'
        elif 'rechazar' in request.POST:
            denuncia.estado = 'rechazada'
        denuncia.save()
        return redirect('lista_denuncias')  # Redirige a la lista de denuncias después de la gestión
    
    return render(request, 'gestiones/denuncias/gestionar_denuncia.html', {'denuncia': denuncia})


@login_required
def crear_publicacion(request):
    if request.method == 'POST':
        form = PublicacionForm(request.POST, request.FILES)  # Agrega request.FILES
        if form.is_valid():
            publicacion = form.save(commit=False)
            publicacion.autor = request.user
            publicacion.save()
            return redirect('home')  # Redirige a la página de inicio después de crear la publicación
    else:
        form = PublicacionForm()

    return render(request, 'gestiones/publicaciones/crear_publicacion.html', {'form': form})

# Comprobar si el usuario pertenece al área de Recursos Humanos
def is_hr_user(user):
    return user.area and user.area.nombre == 'Recursos Humanos'

@login_required
@user_passes_test(is_hr_user)
def subir_documento_empresa(request):
    if request.method == 'POST':
        form = DocumentoEmpresaForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.creado_por = request.user
            documento.save()
            return redirect('lista_documentos_empresa')
    else:
        form = DocumentoEmpresaForm()
    
    return render(request, 'gestiones/documentos/subir_documento.html', {'form': form})


@login_required
def lista_documentos_empresa(request):
    documentos = DocumentoEmpresa.objects.all()
    return render(request, 'gestiones/documentos/lista_documentos.html', {'documentos': documentos})

def descargar_documento_empresa(request, pk):
    documento = get_object_or_404(DocumentoEmpresa, pk=pk)
    
    # Registrar la descarga en la base de datos
    DescargaDocumento.objects.create(documento=documento, usuario=request.user)

    # Lógica para servir el archivo
    response = HttpResponse(documento.archivo, content_type='application/pdf')  # Cambia el tipo si no es PDF
    response['Content-Disposition'] = f'attachment; filename={documento.archivo.name}'
    return response

@login_required
def ver_descargas_documento(request, documento_id):
    documento = get_object_or_404(DocumentoEmpresa, id=documento_id)
    descargas = DescargaDocumento.objects.filter(documento=documento)
    return render(request, 'gestiones/documentos/ver_descargas.html', {'documento': documento, 'descargas': descargas})

@login_required
@user_passes_test(lambda u: u.groups.filter(name='supervisores').exists() and u.area.nombre == "Recursos Humanos")
def eliminar_documento_empresa(request, documento_id):
    documento = get_object_or_404(DocumentoEmpresa, id=documento_id)

    if request.method == 'POST':
        # Elimina el documento si se confirma
        documento.delete()
        messages.success(request, "El documento ha sido eliminado correctamente.")
        return redirect('lista_documentos_empresa')

    # Redirige a la página de confirmación
    return render(request, 'gestiones/documentos/eliminar_documento_confirmacion.html', {'documento': documento})